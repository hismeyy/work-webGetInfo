import os
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter


def excel_export(data):
    current_time = datetime.now()
    # 格式化时间字符串
    formatted_time = current_time.strftime("%Y%m%d%H%M")

    # 创建一个新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = formatted_time

    # 写入表头
    headers = ["标题", "地址", "发布时间"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # 写入数据
    for row_num, item in enumerate(data, start=2):
        ws[f'A{row_num}'] = item['title']
        ws[f'B{row_num}'] = item['href']
        ws[f'C{row_num}'] = item['date']

    # 保存工作簿到文件
    wb.save(
        f"{os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))}\\result\\{formatted_time}.xlsx")

    print(f"数据已成功导出到 {formatted_time}.xlsx")
